import os
import fnmatch
import openpyxl as px
from bs4 import BeautifulSoup
from collections import Counter
from chardet.universaldetector import UniversalDetector
from openpyxl.styles import Border, Side

## -------------------------------------
## ファイルリスト作成
## -------------------------------------

def filelist(root):
    wb = px.load_workbook('./filelist.xlsx') #書き込み対象のExcel
    ws = wb.active
    START_COL = 4 #開始行
    ROW_C = 67 #C列
    ROW_M = 77 #M列
    ROW_V = 86 #V列
    ID_LIST = [65,1,1,1,1,1,1,1,1,1] #65 = A
    PREV_COUNTER = 0
    PREV_DIRECTORY = ''

    ## 現在のディレクトリを再帰的に検索
    for dirpath, dirname, filename in os.walk(root):
        for FILENAME in filename:
            ## 拡張子「.html」or「.php」のファイルに絞る
            if fnmatch.fnmatch(FILENAME, '*.html') or fnmatch.fnmatch(FILENAME, '*.php'):
                ## 対象ファイルパス
                html = os.path.join(dirpath, FILENAME)

                ## 対象ファイルの文字コード判定
                detector = UniversalDetector()
                with open(html, mode='rb') as f:
                    for binary in f:
                        detector.feed(binary)
                        if detector.done:
                            break
                detector.close()

                ## ルートパスは不要なため置換
                PATH = html.replace(root,'')
                DIRECTORY = dirpath.replace(root,'')+'/'
                print(PATH)

                ## html情報を取得
                try:
                    soup = BeautifulSoup(open(html, encoding=detector.result['encoding']), "lxml")

                    ## title情報を取得
                    HEAD = soup.find("head")
                    TITLE_CONTENT = HEAD.find("title")
                    if TITLE_CONTENT != None:
                        TITLE = TITLE_CONTENT.text
                    else:
                        ## titleが無い場合は「null」をセット
                        TITLE = 'null'

                    ## keywords情報を取得
                    META_KEYWORDS = HEAD.find('meta',attrs={'name':'keywords'})
                    if META_KEYWORDS != None:
                        KEYWORDS = META_KEYWORDS.attrs['content']
                    else:
                        ## keywordsが無い場合は「null」をセット
                        KEYWORDS = 'null'

                    ## description情報を取得
                    META_DESCRIPTION = soup.find('meta',attrs={'name':'description'})
                    if META_DESCRIPTION != None:
                        DESCRIPTION = META_DESCRIPTION.attrs['content']
                    else:
                        ## descriptionが無い場合は「null」をセット
                        DESCRIPTION = 'null'

                ## html情報が取得出来なかった場合
                except:
                    TITLE = 'エラー'
                    KEYWORDS = ''
                    DESCRIPTION = ''

                ## 階層を取得
                COUNTER = Counter(PATH)

                ## ページIDをセット
                if PREV_COUNTER >= COUNTER['/']:
                    ID_LIST[(COUNTER['/']-1)] = ID_LIST[(COUNTER['/']-1)]+1
                    for i in range(COUNTER['/'], 6):
                        ID_LIST[i] = 1

                if COUNTER['/'] == 2:
                    if PREV_DIRECTORY != DIRECTORY:
                        ID_LIST[0] = ID_LIST[0]+1
                        ID_LIST[1] = 1

                ## 1つ前のページ情報を保存
                PREV_COUNTER = COUNTER['/']
                PREV_DIRECTORY = DIRECTORY

                ## ページIDを記載
                num = 0
                if COUNTER['/'] == 1:
                    ROW_ID = 2
                else:
                    ROW_ID = COUNTER['/']

                for i in range(ROW_C, ROW_C+ROW_ID):
                    if i == ROW_C:
                        ws[str(chr(i))+str(START_COL)].value = str(chr(ID_LIST[0]))
                    else:
                        ws[str(chr(i))+str(START_COL)].value = ID_LIST[num]
                    num = num+1

                ## タイトルを記載
                for i in range(ROW_M, ROW_V):
                    if i == ROW_M+(COUNTER['/']-1):
                        ws[str(chr(i))+str(START_COL)].value = TITLE

                    ## 自階層の列より左の場合
                    if i - ROW_M < (COUNTER['/']-1):
                        ws[str(chr(i))+str(START_COL)].border = Border(
                        left=Side(style='thin', color='000000'),
                        )
                    ## 自階層の列の場合
                    elif i - ROW_M == (COUNTER['/']-1):
                        ws[str(chr(i))+str(START_COL)].border = Border(
                            top=Side(style='thin', color='000000'),
                            left=Side(style='thin', color='000000'),
                        )
                    ## 自階層の列より右の場合
                    else:
                        ws[str(chr(i))+str(START_COL)].border = Border(
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000'),
                        )

                ## 各ページ情報を記載
                # ws['R'+str(START_COL)].value = DIRECTORY      #ディレクトリ
                # ws['S'+str(START_COL)].value = FILENAME       #ファイル名
                ws['W'+str(START_COL)].value = PATH           #パス
                ws['X'+str(START_COL)].value = KEYWORDS       #keywords
                ws['Y'+str(START_COL)].value = DESCRIPTION    #discription

                START_COL = START_COL+1

    ## 最終行の枠線調整
    for i in range(ROW_M, ROW_V):
        if i == ROW_M:
            ws[str(chr(i))+str(START_COL)].border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
            )
        elif i == ROW_V:
            ws[str(chr(i))+str(START_COL)].border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
            )
        else:
            ws[str(chr(i))+str(START_COL)].border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
            )

    ## Excel保存
    wb.save('./filelist.xlsx')

## 実行（カレントディレクトリ）
if __name__ == '__main__':
    filelist(os.getcwd())
